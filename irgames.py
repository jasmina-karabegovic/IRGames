from email.mime import base
import sys
from turtle import shape

try:
    import numpy as np
except ImportError:
    sys.stderr.write("Error! NumPy python library not detected!! \n")
    sys.exit(1)

try:
    import cdd 
except ImportError:
    sys.stderr.write("Error! python-cdd library not detected!! \n")
    sys.exit(1)   
           
try:
    import matplotlib.pyplot as plt
except ImportError:
    sys.stderr.write("Error! matplotlib.pyplot python library not detected!! \n")
    sys.exit(1)
    
from fractions import Fraction as frac
from scipy.spatial import ConvexHull
from shapely.geometry import Polygon
from openpyxl import load_workbook
import os

try:
    import pathlib
except ImportError:
    sys.stderr.write("Error! pathlib python library not detected!! \n")
    sys.exit(1)

try:
    import sympy as sy
except ImportError:
    sys.stderr.write("Error! SymPy python library not detected!! \n")
    sys.exit(1)

from sympy.abc import x,y
import nashpy as nash
import logging
import argparse
from multiprocessing import Pool
from multiprocessing import cpu_count
import time
from pathlib import Path
from rdp import rdp
import numpy as np
from multiprocessing import cpu_count
from multiprocessing import Pool
import random

parser = argparse.ArgumentParser(description='List of all arguments that can be used.', add_help=False)

parser.add_argument('-h', '--help', action='help', default=argparse.SUPPRESS, help='Show this help message and exit.')

parser.add_argument("-output", metavar='',help="Direct where the output files should be created. If not provided the output path will automatically be set in the location of the current working directory",type=Path)

parser.add_argument("-number_type", help="The decision of how fast and accurate the script should work. The user can opt for float (fast) or fraction (accurate). Default: float.", choices = {"float","fraction"},default="float")

parser.add_argument('-difference',metavar='', type=float, default=0.005, help='Pre-defined error bound between polytope areas. Default = 0.005.')

parser.add_argument("-rounds",metavar='', help="Enter the number of rounds you would like the algorithm to compute. Default: 2.", type=int, default=2)

parser.add_argument('-mp',metavar='', type=int, help="This argument specifies number of processes for parallel processing (parallel computing). If opted to use maximum amount of computational resources on the machine, type '--processes 0'. If disabled, the algorithm is using a single process.")

parser.add_argument("-output_vertices", action="store_true", help="Saving output in the text format in the Output_vertices folder; increases computation time, but makes the availability of the output after the script finishes. If it is off, the tool will only save the final round. Default: off. Call -ov or --output_vertices to turn it on.") 

parser.add_argument("-s_from",metavar='',help="Continuation of previous rounds. Input as an integer what round you wish to continue the analysis from.",type=int)

parser.add_argument("-log", action="store_true", help="Log of activities: the payoff matrix, probability distribution of the signals, the discount factor delta, number of rounds, computational time needed  (depends on values that have been used: float or fraction), etc. Default: off. Call -l or --log to turn on.")

parser.add_argument("-plot", metavar='',help="If chosen gives the possibility to the user to plot only one round. Input as an integer which round the user wants to visualize.",type=int)

parser.add_argument("-sp", metavar = '', type = float, help="Simplifying polytopes using the Ramer-Douglas-Peucker line simplification algorithm. Suggested epsilon value = 0.000001.")

parser.add_argument("-verify", action = "store_true", help="Working example to check for correct installations of packages. Default: off. Call -verify to enable it.")

args = parser.parse_args()

plt.plot()


num_processes=args.mp 
if num_processes==None:
    num_processes=1
if num_processes==0:
    num_processes=cpu_count()

logger = logging.getLogger(__name__)
if args.log:
    logging.basicConfig(filename='logfile.log',level=logging.DEBUG,format="%(asctime)s %(message)s", filemode="w")
    logging.getLogger('matplotlib.font_manager').disabled = True
else:
    logging.basicConfig(format="%(levelname)s: %(message)s")

#the path where the output folders will be created; the output will be saved as an image and as a txt file. 
if args.output == None or args.output == 0:
    path_cwd=pathlib.Path.cwd()
else:
    path_cwd=args.output

png_file = f'{path_cwd}/Output_png'
isExist_png = os.path.exists(png_file)
if args.output_vertices:
    txt_file = f'{path_cwd}/Output_vertices'
    isExist_txt = os.path.exists(txt_file)
    if not isExist_txt:
        os.makedirs(txt_file)
try:
    if not isExist_png:
        os.makedirs(png_file)
except OSError as err:
    print(err)

cont_file = f'{path_cwd}/Output_halfspaces'
isExist_cont = os.path.exists(cont_file)
if not isExist_cont:
    os.makedirs(cont_file)

poly_area = f"{path_cwd}/poly_area.txt"

if args.plot==None:
    pass
elif args.plot==0:
    print("Error invalid input. Round 0 does not exist")
    exit()
else:
    plotting=np.loadtxt(f"{path_cwd}/Output_vertices/Extreme_points_iteration_{args.plot}.csv",delimiter=",")
    plotting_hull = ConvexHull(plotting)
    plt.plot(plotting[:,0], plotting[:,1], 'o')
    for simplex in plotting_hull.simplices:
        plt.plot(plotting[simplex, 0], plotting[simplex, 1], 'k-')
    plt.title(f"Iteration_Number_{args.plot}")
    plt.savefig(f"{path_cwd}/Iteration_Number_{args.plot}")
    print(f"Png saved as *Iteration_Number_{args.plot}*")
    exit()

if args.verify:
    logging.info(f"Selected input option: {args.verify}, for verification of instalation")
    number_of_actions_1 = 2
    number_of_actions_2 = 2
    number_of_signals = 2

    payoff_player_1 = np.array([[2, -1], [3, 0]])
    payoff_player_2 = np.array([[2, 3], [-1, 0]])

    reshaped_pm = np.array([[[frac(2,3),frac(1,3)],[frac(1,2), frac(1,2)]],
                    [[frac(1,2), frac(1,2)], [frac(1,4), frac(3,4)]]])
    delta = frac(9,10)
    stage_game=np.array([[[2,2],[-1,3]],[[3,-1],[0,0]]])

else:
    #load_excel file.
    excel_file = load_workbook(filename="input_data.xlsm")
    if excel_file == None:
        excel_file = load_workbook(filename="input_data.xlsx")
    if excel_file == None:
        print("Can't find the excel file. Please make sure to create one in your working directory.")
    
    data_sheet=excel_file["Raw_data"]
    #load_values: values retrieved from the excel file. 
    number_of_actions_1=data_sheet.cell(row=2, column=2).value
    number_of_actions_2=data_sheet.cell(row=3, column=2).value
    number_of_signals=data_sheet.cell(row=4, column=2).value
    delta=data_sheet.cell(row=5, column=2).value

    row_counter_p1=3
    payoff_player_1=[]
    temp_array=[]

    for i in range(number_of_actions_1):
        temp_array=[]
        for j in range(number_of_actions_1):
            temp_value=data_sheet.cell(row=row_counter_p1, column=6).value
            temp_array.append(temp_value)
            row_counter_p1=row_counter_p1+1
        payoff_player_1.append(temp_array) 
    payoff_player_1=np.asarray(payoff_player_1)

    row_counter_p2=3
    payoff_player_2=[]
    temp_array=[]

    for i in range(number_of_actions_2):
        temp_array=[]
        for j in range(number_of_actions_2):
            temp_value=data_sheet.cell(row=row_counter_p2, column=7).value
            temp_array.append(temp_value)
            row_counter_p2=row_counter_p2+1
        payoff_player_2.append(temp_array) 
    payoff_player_2=np.asarray(payoff_player_2)


    row_counter_pm=3
    reshaped_pm=[]
    for k in range (number_of_actions_1):
        temp_var=[]
        for i in range(number_of_actions_1):
            col_counter=11
            payoff_array=[]
            for j in range(number_of_signals):
                temp_value=data_sheet.cell(row=row_counter_pm, column=col_counter).value
                payoff_array.append(temp_value)
                col_counter=col_counter+1
            row_counter_pm=row_counter_pm+1
            temp_var.append(payoff_array)
        reshaped_pm.append(temp_var)
    reshaped_pm=np.array(reshaped_pm)


    row_counter_sg=3
    stage_game=[]
    for k in range (number_of_actions_1):
        temp_var=[]
        for i in range(number_of_actions_1):
            col_counter_sg=6
            payoff_array=[]
            for j in range(2):
                temp_value=data_sheet.cell(row=row_counter_sg, column=col_counter_sg).value
                payoff_array.append(temp_value)
                col_counter_sg=col_counter_sg+1
            row_counter_sg=row_counter_sg+1
            temp_var.append(payoff_array)
        stage_game.append(temp_var)
    stage_game=np.array(stage_game)
    
number_of_action_profiles = number_of_actions_1*number_of_actions_2
number_of_outcomes=number_of_action_profiles*number_of_signals

# slightly adjusted from https://github.com/drvinceknight/nashpy to compute multiple equilibria
def select_any_nash_equilibrium(payoffs1, payoffs2): 
    # Create a Nash game from the payoffs
    nash_game = nash.Game(payoffs1, payoffs2)
    
    # Find all Nash equilibria
    equilibria = list(nash_game.support_enumeration())
    
    # Raise an error if no equilibria found
    if not equilibria:
        raise ValueError("No Nash Equilibrium found.")
    
    # Select any one Nash equilibrium (e.g., randomly)
    selected_equilibrium = random.choice(equilibria)
    
    # Extract strategies for each player
    player1_strategy, player2_strategy = selected_equilibrium
    
    # Calculate the equilibrium payoffs for both players
    expected_payoff_player_1 = nash_game[player1_strategy, player2_strategy][0]
    expected_payoff_player_2 = nash_game[player1_strategy, player2_strategy][1]
    
    # Combine both payoffs into a single array and return
    return np.array([expected_payoff_player_1, expected_payoff_player_2])

# Using the function to get an equilibrium payoff
Nash_Equilibrium = select_any_nash_equilibrium(payoff_player_1, payoff_player_2)
difference = args.difference
if args.log:
    logging.info(f"Value Type: {args.number_type}")
    logging.info(f"Number of processes used: {num_processes}")
    logging.info(f"Polygon error bound: {difference}")
    logging.info(f"Number of actions player 1: {number_of_actions_1}")
    logging.info(f"Number of actions player 2: {number_of_actions_2}")
    logging.info(f"Number of signals: {number_of_signals}")
    logging.info(f"Signal structure:\n {reshaped_pm}")
    logging.info(f"Stage game:\n {stage_game}")
    logging.info(f"Delta: {delta}")
    if args.sp:
        logging.info(f"Simplifying Polytopes(epsilon value): {format(args.sp,'f')} \n")

    logging.info("Script calculated values")

    
    #to compute polytope vertices and halfspaces, the definitions have been used from: https://github.com/stephane-caron/pypoman/blob/main/pypoman/duality.py
def compute_poly_vertices(A, b):
    b = b.reshape((b.shape[0], 1))
    mat = cdd.Matrix(np.hstack([b, -A]), number_type=args.number_type)
    mat.rep_type = cdd.RepType.INEQUALITY
    P = cdd.Polyhedron(mat)
    g = P.get_generators()
    V = np.array(g)
    vertices = []
    for i in range(V.shape[0]):
        if V[i, 0] != 1: continue
        if i not in g.lin_set:
            vertices.append(V[i, 1:])
    return vertices

def compute_polytope_halfspaces(vertices):
    V = np.vstack(vertices)
    t = np.ones((V.shape[0], 1))  # first column is 1 for vertices
    tV = np.hstack([t, V])
    mat = cdd.Matrix(tV, number_type=args.number_type)
    mat.rep_type = cdd.RepType.GENERATOR
    P = cdd.Polyhedron(mat)
    bA = np.array(P.get_inequalities())
    if bA.shape == (0,):  # bA == []
        return bA
    # the polyhedron is given by b + A x >= 0 where bA = [b|A]
    b, A = np.array(bA[:, 0]), -np.array(bA[:, 1:])
    return (A, b)

#symbols as continuation payoffs. 
def create_symbols_and_payoffs(num_signals):
    # Create a list of alphabet symbols starting from 'p'
    symbols = [chr(112 + i) for i in range(num_signals)]  #This creates a list of symbols starting from 'p' and continues alphabetically. The chr function converts ASCII values to characters, and 112 is the ASCII value for 'p'.
    # Create sympy symbols
    sympy_symbols = sy.symbols(" ".join(symbols), real=True)
    # Create the payoff array
    cont_payoffs = np.array(sympy_symbols)
    return sympy_symbols, cont_payoffs

sympy_symbols, cont_payoffs = create_symbols_and_payoffs(number_of_signals)


#Please note that the naming scheme has been adjusted for the working example from my original paper on the role of international courts. One could think of it as if we are analysing the prisoners' dilemma. The aim in the following section is to transform the set from its minimal extreme point represenetation to its half-space representation, including the payoff set itself and the incentive constraints.

#computing the incentive constraints for player 1
incentives_P1=[]
for i in range(number_of_actions_1):
    for j in range(number_of_actions_2):
        new_equation = stage_game[i][j][0]*(1-delta) + delta*(np.dot(reshaped_pm[i][j], np.transpose(cont_payoffs)))
        incentives_P1.append(new_equation)
        
        
#computing the incentive constraints for player 2
incentives_P2=[]
for i in range(number_of_actions_2):
    for j in range(number_of_actions_1):
        new_equation = stage_game[i][j][1]*(1-delta)+delta*(np.dot(reshaped_pm[i][j],np.transpose(cont_payoffs)))
        incentives_P2.append(new_equation)


if number_of_actions_1==2:
 #conditions needed to be satisfied for player 1 to cooperate 
    incentives_cooperate_P1=[]
    for i in range(number_of_actions_1):
        incentive_list=incentives_P1[i]-incentives_P1[i+2]
        incentives_cooperate_P1.append(incentive_list)

    #conditions needed to be satisfied for player 1 to defect 
    incentives_defect_P1=[]
    for i in range(number_of_actions_1):
        incentive_list=incentives_P1[i+2]-incentives_P1[i]
        incentives_defect_P1.append(incentive_list)    
    ICC_P1=np.vstack((incentives_cooperate_P1,incentives_defect_P1)) #merging them together
    #the same is repeated if the number of actions is 3. 
elif number_of_actions_1==3:

    incentives_cooperate_P1=[]
    for i in range(number_of_actions_1):
        counter=i+3
        for j in range(number_of_actions_1-1):
            incentive_list=incentives_P1[i]-incentives_P1[counter]
            counter=counter+3
            incentives_cooperate_P1.append(incentive_list)
    incentives_cd_P1=[]
    for i in range(number_of_actions_1):
        counter=i
        for j in range(number_of_actions_1-1):
            incentive_list=incentives_P1[i+3]-incentives_P1[counter]
            counter=counter+6
            incentives_cd_P1.append(incentive_list)
    incentives_defect_P1=[]
    for i in range(number_of_actions_1):
        counter=i
        for j in range(number_of_actions_1-1):
            incentive_list=incentives_P1[i+6]-incentives_P1[counter]
            counter=counter+3
            incentives_defect_P1.append(incentive_list)
    ICC_P1=np.vstack((incentives_cooperate_P1,incentives_cd_P1,incentives_defect_P1)) #merging them together
else:
    print("Warning number of actions outside of the range")
    sys.exit()
    
if number_of_actions_2==2:
     #conditions needed to be satisfied for player 2 to cooperate 
    incentives_cooperate_P2=[]
    counter=0
    for i in range(number_of_actions_2):
        incentive_list=incentives_P2[counter]-incentives_P2[counter+1]
        incentives_cooperate_P2.append(incentive_list)
        counter=2

    #conditions needed to be satisfied for player 2 to defect
    counter=0
    incentives_defect_P2=[]
    for i in range(number_of_actions_2):
        incentive_list=incentives_P2[counter+1]-incentives_P2[counter]
        incentives_defect_P2.append(incentive_list)
        counter=2 

    icc = []
    for i in range(number_of_actions_2):
        element = np.vstack([[incentives_cooperate_P2[i], incentives_defect_P2[i]]])
        icc.append(element)
    ICC_P2 = np.vstack((icc))
    
    #the same is repeated if player 2 has 3 actions.
elif number_of_actions_2==3:
    incentives_cooperate_P2=[]
    for i in range(3):
        for j in range(3):
            if i != j:
                incentive_list=incentives_P2[i]-incentives_P2[j]
                incentives_cooperate_P2.append(incentive_list)

    incentives_cd_P2=[]
    for i in range(3):
        for j in range(3):
            if i != j:
                incentive_list=incentives_P2[i+3]-incentives_P2[j+3]
                incentives_cd_P2.append(incentive_list)


    incentives_defect_P2=[]
    for i in range(3):
        for j in range(3):
            if i != j:
                incentive_list=incentives_P2[i+6]-incentives_P2[j+6]
                incentives_defect_P2.append(incentive_list)

    icc = []
    for i in range(number_of_actions_1+number_of_actions_2): 
        element = np.vstack([[incentives_cooperate_P2[i],incentives_cd_P2[i], incentives_defect_P2[i]]])
        icc.append(element)
    ICC_P2 = np.vstack((icc))

    #the incentive constraints can always be checked if logging is enabled. 
if args.log:
    logging.info(f"Incentive constrains player 1:{incentives_P1}")
    logging.info(f"Incentive constrains player 2:{incentives_P2} \n")

    #to see if there are any continuation payoffs, we are checking for the generators in the incentive constraints we calculated. 
has_generators_P1=any(term.has(sy.Symbol)for term in ICC_P1.flat)
has_generators_P2=any(term.has(sy.Symbol)for term in ICC_P2.flat)

if has_generators_P1 == False or has_generators_P2 == False:
    print("Players play Nash equilibrium strategies")
    if args.log:
        logging.info(f"Assumption: Both players' incentives have generators.")
        logging.info(f"Generators P1: {has_generators_P1}")
        logging.info(f"Generators P2: {has_generators_P2}")
    exit()

#extracting coefficients from the polynomials defining the incentive constraints, and putting them into arrays. We do that for both players.
poly_list_P1=[]
poly_coeff_P1=[]
if number_of_actions_1==2:
    loop_lenght=number_of_actions_1
elif number_of_actions_1==3:
    loop_lenght= number_of_actions_1 + number_of_actions_2
    
for i in range(number_of_actions_1):
    for j in range(loop_lenght):
        polynomial=sy.poly(ICC_P1[i][j],sympy_symbols)
        coefficient=polynomial.coeffs()
        coefficients = [polynomial.coeff_monomial(symbol) for symbol in sympy_symbols]
        coefficients.append(polynomial.coeff_monomial(1))  # For the constant term
        poly_list_P1.append(polynomial)
        poly_coeff_P1.append(coefficients)

poly_list_P2=[]
poly_coeff_P2=[]
for i in range(loop_lenght):
    for j in range(number_of_actions_2):
        polynomial=sy.poly(ICC_P2[i][j],sympy_symbols)
        coefficients = [polynomial.coeff_monomial(symbol) for symbol in sympy_symbols]
        coefficients.append(polynomial.coeff_monomial(1))  # For the constant term
        poly_list_P2.append(polynomial)
        poly_coeff_P2.append(coefficients)

coeff_matrix_P1=np.reshape(poly_coeff_P1,(int(len(poly_coeff_P1)),int(len(max(poly_coeff_P1,key=len)))))
coeff_matrix_P2=np.reshape(poly_coeff_P2,(int(len(poly_coeff_P2)),int(len(max(poly_coeff_P2,key=len)))))
        
M_P1 = coeff_matrix_P1[:, 0:(int(len(max(poly_coeff_P1,key=len)))-1)]#creating matrices that go with the continuation payoffs for player 1; that is, the coefficient matrix. 
M_P2 = coeff_matrix_P2[:, 0:(int(len(max(poly_coeff_P2,key=len)))-1)]#creating matrices that go with the continuation payoffs for player 2; that is, the coefficient matrix. 
b_P1 = coeff_matrix_P1[:,-1] #extracting the b coefficient so the following holds: Ax ≤ b.
b_P2= coeff_matrix_P2[:,-1] #extracting the b coefficient so the following holds: Ax ≤ b.

if number_of_actions_1==2: #considering the case if players have two actions
    extended_M_P1=np.pad(M_P1,((0,0),(0,number_of_signals)),mode="constant",constant_values=0) #extending the arrays for the number of signals 
    extended_M_P2=np.pad(M_P2,((0,0),(number_of_signals,0)),mode="constant",constant_values=0)

    extended_M_P1_ = np.multiply(extended_M_P1, -1) #transforming the coefficients so the Ax ≤ b form holds
    extended_M_P2_ = np.multiply(extended_M_P2, -1)
    incentive_constraints=[]
    b_vector=[]
    for i in range(number_of_action_profiles):
        temp_incentive_constraints=np.vstack([extended_M_P1_[i],extended_M_P2_[i]])
        temp_b_vector=np.vstack([b_P1[i],b_P2[i]])
        incentive_constraints.append(temp_incentive_constraints)
        b_vector.append(temp_b_vector)
    incentive_constraints=np.asarray(incentive_constraints) #pairing the incentive constraints for the given action profiles

    b_vector=np.asarray(b_vector) #pairing the b_vector for every action profile
     
elif number_of_actions_1==3:  #considering the case if players have three actions
    counter=0
    M_P1_paired=[]
    for i in range(int(len(M_P1)/2)):
        temp_list_pairs_P1=np.vstack((M_P1[counter],M_P1[counter+1]))
        M_P1_paired.append(temp_list_pairs_P1)
        counter=counter+2
    M_P1_paired=np.asarray(M_P1_paired)

    counter_i=0
    counter_j=0
    M_P2_paired=[]
    for i in range(3):
        for j in range(3):
            temp_list_pairs_P2=np.vstack((M_P2[counter_i],M_P2[counter_j+3]))
            M_P2_paired.append(temp_list_pairs_P2)
            counter_i=counter_i+6
            counter_j=counter_j+6
        counter_i=i+1
        counter_j=i+1
    M_P2_paired=np.asarray(M_P2_paired)

    extended_M_P1=np.pad(M_P1_paired,((0,0),(0,0),(0,number_of_signals)),mode="constant",constant_values=0) #extending the arrays for the number of signals
    extended_M_P2=np.pad(M_P2_paired,((0,0),(0,0),(number_of_signals,0)),mode="constant",constant_values=0)

    extended_M_P1_ = np.multiply(extended_M_P1, -1) #transforming the coefficients so the Ax ≤ b form holds
    extended_M_P2_ = np.multiply(extended_M_P2, -1)
  
    counter=0
    b_P1_paired=[]
    for i in range(int(len(M_P1)/2)):
        temp_list_pairs_b_P1=np.vstack((b_P1[counter],b_P1[counter+1]))
        b_P1_paired.append(temp_list_pairs_b_P1)
        counter=counter+2
    b_P1_paired=np.asarray(b_P1_paired) 

    counter_i=0
    counter_j=0
    b_P2_paired=[]
    for i in range(3):
        for j in range(3):
            temp_list_pairs_b_P2=np.vstack((b_P2[counter_i],b_P2[counter_j+3]))
            b_P2_paired.append(temp_list_pairs_b_P2)
            counter_i=counter_i+6
            counter_j=counter_j+6
        counter_i=i+1
        counter_j=i+1
    b_P2_paired=np.asarray(b_P2_paired)
    b_vector=[]
    for i in range (0 , number_of_action_profiles):
        temp_b_vector=np.vstack((b_P1_paired[i],b_P2_paired[i]))
        b_vector.append(temp_b_vector)
    b_vector=np.asarray(b_vector) #pairing the b_vectors for each action profile

    incentive_constraints=[] #pairing incentive constraints for every action profile
    for i in range(number_of_action_profiles):
        temp_incentive_constraints=np.vstack([extended_M_P1_[i],extended_M_P2_[i]])
        incentive_constraints.append(temp_incentive_constraints)
    incentive_constraints=np.asarray(incentive_constraints)
else:
    print("Warning")
    sys.exit()


individual_probabilities=[]
for i in range(number_of_actions_1):
    for j in range(number_of_actions_1):
        y=np.vstack([reshaped_pm[i][j],reshaped_pm[i][j]])
        individual_probabilities.append(y)
individual_probabilities=np.asarray(individual_probabilities)
zero_array=np.zeros(number_of_signals,dtype=int) #extending the signal structure over all possible signals

extended_individual_probabilities=[]
for i in range(number_of_action_profiles):
    for j in range(2):
        if (j % 2) == 0:
            temp_extended_individual_probabilities=np.hstack([individual_probabilities[i][j],zero_array]) #added based on number of signals
        
        else:
            temp_extended_individual_probabilities=np.hstack([zero_array,individual_probabilities[i][j]])
        extended_individual_probabilities.append(temp_extended_individual_probabilities)
extended_individual_probabilities=np.asarray(extended_individual_probabilities) 


def create_signals(num_signals):
    signal_length = 2 * num_signals  # The length of each signal array
    signals = []

    for i in range(num_signals):
        signal = np.zeros((2, signal_length), dtype=int)  # Specify dtype=int to create integer arrays
        signal[0, i] = 1
        signal[1, num_signals + i] = 1
        signals.append(signal)
    
    return signals

signals = create_signals(number_of_signals) #creating arrays to expand the signal structures

action_profile_probabilities=[]
counter=0
for i in range(number_of_action_profiles):
    temp_action_profile_probabilities=np.vstack([extended_individual_probabilities[counter],extended_individual_probabilities[counter+1]])
    action_profile_probabilities.append(temp_action_profile_probabilities)
    counter=counter+2
action_profile_probabilities=np.asarray(action_profile_probabilities) #pairing these probabilities over all action profiles 

#minmax for player 1
max_values_1=np.max(payoff_player_1,axis=0)
minmax_1=np.min(max_values_1)

#minmax for player 2
max_values_2=np.max(payoff_player_2,axis=1)
minmax_2=np.min(max_values_2)


logging.info(f"MinMax Player1: {minmax_1}") #if logging is enabled, this will be part of the logged information
logging.info(f"MinMax Player2: {minmax_2}")

n_initial=np.hstack([minmax_1,minmax_2]) #excluding the part of the feasible set that is not included in the individually rational feasible payoff set
n=np.hstack([n_initial] * number_of_signals)
n_negative = np.multiply(n,-1) #making sure it is in the Ax ≤ b form

N = np.vstack([-np.eye(int(number_of_signals*2))]) #making sure that the continuation payoff are at least the minmax payoff
base_game = []
for i in range(number_of_actions_1):
    for j in range(number_of_actions_2):
        payoff = stage_game[i][j]
        base_game.append(payoff)

base_game = np.array(base_game)  

new_minmax = np.array([[-1,0],
                      [0, -1]])
new_minmax_1 = np.array([-minmax_1,-minmax_2])   #including the minmax payoffs in the H representaion

A,b = compute_polytope_halfspaces(base_game)

A_a = np.concatenate((A, new_minmax), axis = 0)
b_b = np.concatenate((b, new_minmax_1), axis = 0)

base_game_plot = np.array(compute_poly_vertices(A_a, b_b))  #computing the vertices of the individually rational feasible payoff set i.e., representing it in its minimal extreme points form

#print(base_game_plot)
hull = ConvexHull(base_game_plot)
plt.plot(base_game_plot[:,0], base_game_plot[:,1], 'o')
for simplex in hull.simplices:
    plt.plot(base_game_plot[simplex, 0], base_game_plot[simplex, 1], 'k-') #plot the individually rational feasible payoff set
#plt.show()

if args.s_from==None or args.s_from==0: #if we are computing everything from round 0
    A_new = np.concatenate([np.dot(A, signal) for signal in signals], axis=0) #concatenating the signals and the b vectors so all elements are included in the H representation of the set 
    b_new = np.concatenate([b] * number_of_signals, axis=0) 
    continuation_number = 0
    previous_area_of_polygon = 0
    empty_array = np.array([])
    # check if the file exists
    if os.path.exists(poly_area):
        os.remove(poly_area)
else:
    A_new=np.loadtxt(f"{path_cwd}/Output_halfspaces/A_new{args.s_from}.csv",delimiter=",") #load a previously finished round and upload it in its H-representation
    b_new=np.loadtxt(f"{path_cwd}/Output_halfspaces/b_new{args.s_from}.csv",delimiter=",")
    continuation_number=int(args.s_from)
    logging.info(f"The log file of the extended use. Round: {continuation_number}")
    with open(poly_area,"r") as file_for_poly_area:
        lines=file_for_poly_area.readlines()
    for line in lines:
        columns=line.strip().split()
        if int(columns[0])== continuation_number:
            previous_area_of_polygon=float(columns[1])
            break
#single core run without multiprocessing
def single_core_main_run(incentive_constraints, A_new, b_vector, b_new, base_game, delta, action_profile_probabilities):
 
    average_payoffs = []
    temp_variable_x=[]
    empty_array = []
    counter=0
    for i in range(len(incentive_constraints)):
        action_profile_icc = incentive_constraints[i]
        A = np.concatenate([action_profile_icc,A_new,N])
        b_vector_action_profile = b_vector[i]
        flat_b_action_profile = b_vector_action_profile.flatten()
        b = np.concatenate([flat_b_action_profile, b_new,n_negative])
        points = np.array(compute_poly_vertices(A,b)) 
        if np.array_equal(empty_array, points):
            print(f" action profile: {i}", "has no continuation payoffs")
            continue
        else:
            temp_var= np.empty(((len(points)),(int(len(points[0])/number_of_signals))))
            for k in range(len(points)):
                temp_var[k,:] = np.multiply(base_game[i],(1-delta)) +  delta * (np.dot(action_profile_probabilities[i],points[k]))
            average_payoffs.append(temp_var)

            for i in range(len(average_payoffs)):
                average_payoffs[i] = np.maximum(average_payoffs[i], minmax_1)

            average_payoffs_f=np.asarray(average_payoffs,dtype=object)
            average_payoffs_f=np.concatenate((average_payoffs_f), axis = 0)
        counter=counter+1
    temp_variable_x=average_payoffs_f
    return temp_variable_x

#calling the function convex_polytope, to convexify the union of sets
def convex_polytope():
    hull_w_n1 = ConvexHull(temp_variable_x)
    index = hull_w_n1.vertices
    epoints = temp_variable_x[index]
    return epoints
#stopping argument of the algorithm
def area_exit(epoints, previous_area_of_polygon, difference):
    polygon = Polygon(epoints)
    poly = polygon.area
    if poly == 0:
        print("The only feasible point is the Nash equilibrium of the game")
        exit()
    diff = abs(previous_area_of_polygon - poly)
    pct_diff = diff / previous_area_of_polygon * 100 if previous_area_of_polygon != 0 else 100
    pct_str = "{:.2f}%".format(pct_diff)
    if diff >= difference:
        logging.info(f"Absolute difference between two polygons: {round(diff, 5)} ({pct_str})")
        previous_area_of_polygon=poly
        if args.s_from:
            with open(poly_area, 'a') as file:
                file.write(f"{iteration_number}\t{previous_area_of_polygon}\tstarted from {args.s_from}\n")
        else:
            with open(poly_area, 'a') as file:
                file.write(f"{iteration_number}\t{previous_area_of_polygon}\n")            
        return previous_area_of_polygon, False  # Indicate to continue the loop
    else:
        logging.info(f"Absolute difference between two polygons is less than the set threshold of {difference}")
        logging.info(f"Final area of the polygon: {round(poly, 5)}")
        print(f"\nAbsolute difference between two polygons is less than the set threshold of {difference}")
        print(f"Stopping the calculations.Final area of the polygon: {round(poly, 5)}\n")
        return previous_area_of_polygon, True  # Indicate to stop the loop
#plotting differnt rounds
def plotting_ ():
    plotting_points=ConvexHull(red_epoints) 
    plt.plot(red_epoints[:,0], red_epoints[:,1], 'o')
    for simplex in plotting_points.simplices:
        plt.plot(red_epoints[simplex,0], red_epoints[simplex,1], 'k-')
    #plt.show()
                                
    plt.title(f"Round {iteration_number}")
    plt.savefig(f"{path_cwd}/Output_png/Iteration_{iteration_number}_{args.number_type}")
    
    if args.output_vertices:
        np.savetxt(f"{path_cwd}/Output_vertices/Extreme_points_iteration_{iteration_number}.csv",red_epoints,delimiter=",") #think about making this optional and maybe even moving it just for the last one
        np.savetxt(f"{path_cwd}/Output_halfspaces/A_new{iteration_number}.csv",A_new,delimiter=",") #think about making this optional and maybe even moving it just for the last one
        np.savetxt(f"{path_cwd}/Output_halfspaces/b_new{iteration_number}.csv",b_new,delimiter=",") #think about making this optional and maybe even moving it just for the last one  
payoffs =[]

#multiprocessing is introduced
def compute_poly_vertices_parallel(args):
    i, action_profile_icc, A_new, b_vector_action_profile, b_new, base_game_i, delta, action_profile_probabilities_i, N, n_negative, minmax_1 = args
    empty_array = []
    A = np.concatenate([action_profile_icc, A_new, N])
    flat_b_action_profile = b_vector_action_profile.flatten()
    b = np.concatenate([flat_b_action_profile, b_new, n_negative])
    points = np.array(compute_poly_vertices(A, b))
    if np.array_equal(empty_array, points):
        print(f" action profile: {i}", "has no continuation payoffs")
        return None
    else:
        temp_var = np.empty(((len(points)), (int(len(points[0]) / number_of_signals)))) #number_of_signals
        for k in range(len(points)):
            temp_var[k, :] = np.multiply(base_game_i, (1 - delta)) + delta * (np.dot(action_profile_probabilities_i, points[k]))
        temp_var = np.maximum(temp_var, minmax_1)
        return temp_var

def parallel_computing(incentive_constraints, A_new, b_vector, b_new, base_game, delta, action_profile_probabilities, num_processes, N, n_negative, minmax_1):
    args = [(i, incentive_constraints[i], A_new, b_vector[i], b_new, base_game[i], delta, action_profile_probabilities[i], N, n_negative, minmax_1) for i in range(len(incentive_constraints))]
    
    with Pool(num_processes) as pool:
        average_payoffs = pool.map(compute_poly_vertices_parallel, args)
    
    # Remove any None results from failed computations
    average_payoffs = [ap for ap in average_payoffs if ap is not None]

    # If there are any valid payoffs, concatenate them, otherwise return an empty array
    if average_payoffs:
        average_payoffs_f = np.concatenate(average_payoffs, axis=0)
    else:
        average_payoffs_f = np.array([])
    
    return average_payoffs_f


if __name__=="__main__":
    start_total=time.perf_counter()
    for j in range(0,args.rounds):
        print(f"Round: {j+1}")
        temp_variable_x=None
        iteration_number = j + continuation_number + 1
        if num_processes==1: #single core computation, no multiprocessing
            temp_variable_x=single_core_main_run(incentive_constraints, A_new, b_vector, b_new, base_game, delta, action_profile_probabilities)
        else:
            temp_variable_x = parallel_computing(incentive_constraints, A_new, b_vector, b_new, base_game, delta, action_profile_probabilities, num_processes, N, n_negative, minmax_1) #multiprocessing enabled
        if np.allclose(np.mean(temp_variable_x), Nash_Equilibrium, atol=0.001):
                print("The set of feseable payoffs is Nash eqilibrium")
                break
        else:
            epoints=convex_polytope()
            if args.sp:
                red_epoints = rdp(epoints, epsilon = args.sp)
            else:
                red_epoints=epoints
            A_1, b_1 = compute_polytope_halfspaces(red_epoints)
            A_new = np.concatenate([np.dot(A_1, signal) for signal in signals], axis=0)
            b_new = np.concatenate([b_1] * number_of_signals, axis=0) 
            plotting_()
            previous_area_of_polygon, should_stop = area_exit(red_epoints, previous_area_of_polygon, difference)
            if should_stop:
                break
    try:
        if args.output_vertices==False:
            np.savetxt(f"{path_cwd}/Final_round",red_epoints,delimiter=",") 
        if num_processes!=1:
            logging.info(f"Implementation: Multiprocessing (amount of processes used: {num_processes}). With {args.number_type} as the argument")
            print(f"Implementation: Multiprocessing (amount of processes used: {num_processes}). With {args.number_type} as the argument")
        elif args.mp==1:
            logging.info(f"Implementation: Single Core. With {args.number_type} as the argument")
            print("Warning: You have set the number of processes to 1. This is the same as running single core and will not benefit from multiprocessing. Defaulting back to Single core")
            print(f"Implementation: Single Core. With {args.number_type} as the argument")
        else:
            logging.info(f"Implementation: Single Core. With {args.number_type} as the argument")
            print(f"Implementation: Single Core. With {args.number_type} as the argument")            
        end_total=time.perf_counter()
    except Exception as e:
        print("An error occurred during saving/logging:", e)
print("Total run time:",round((end_total-start_total),2))
